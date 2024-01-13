import openpyxl
import xlwings as xw


def fix_and_load_excel(file_path, sheet_name):
    try:
        app = xw.App(visible=False)  # Open Excel in the background
        workbook = app.books.open(file_path)
        workbook.save()
        workbook.close()
        app.quit()
        repaired_workbook = openpyxl.load_workbook(file_path)
        print("Workbook repaired and loaded successfully.")

        sheet = repaired_workbook[sheet_name]

        # It clear sheet sheet name remain but content completly removed
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.value = None
        
        # It corroupt drawing.xml and result in file and data loss
        if sheet_name in repaired_workbook.sheetnames:
            del repaired_workbook[sheet_name]
        repaired_workbook.save(file_path)
        print(f"The sheet '{sheet_name}' has been deleted.")

    except Exception as e:
        print(f"Error: {e}")


if __name__ == "__main__":
    file_path = r'E:\Projects\Python\Delete Sheet\1.xlsx'
    fix_and_load_excel(file_path, "Evaluation Warning")
